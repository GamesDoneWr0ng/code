import openpyxl
import os
import xlwings as xw
from PIL import Image
from openpyxl.utils.cell import _get_column_letter
from numba import jit
import time

SIZE = (480, 360)
FPS = 12

STRING_COL_CACHE = {}
for i in range(1, SIZE[0]+1):
    col = _get_column_letter(i)
    STRING_COL_CACHE[i] = col

lastFrame = time.time()
os.chdir("python/Apple/Frames")
frames = sorted(i if len(i) == 8 else "0" + i for i in os.listdir('.'))[1:]
wb = openpyxl.load_workbook("/Users/askborgen/Desktop/code/python/Apple/Meanwhile.xlsx")
xwwb = xw.books.active
update = xwwb.macro("Update")
reSize = xwwb.macro("ReSize")
sheet = wb.active

reSize()

@jit(parallel=True, forceobj=True, target_backend='cuda')
def display(img, lastImg, sh, cache):
    for x in range(SIZE[0]):
        for y in range(SIZE[1]):
            color = "%02x%02x%02x" % img.getpixel((x,y))
            if color == "%02x%02x%02x" % lastImg.getpixel((x,y)):
                continue
            style = openpyxl.styles.colors.Color(rgb=color)
            fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=style)
            sh[f"{cache[(x+1)]}{y+1}"].fill = fill

# animate
lastImg = Image.open(frames[0][1:])
lastImg.mode
for frame in frames[::3]:
    img = Image.open(frame if frame[0] != "0" else frame[1:])
    img.mode

    display(img, lastImg, sheet, STRING_COL_CACHE)
    lastImg = img

    wb.save("/Users/askborgen/Desktop/code/python/Apple/Meanwhile.xlsx")
    update()
    print(frame, 1 / (lastFrame - time.time()), "fps")
    lastFrame = time.time()